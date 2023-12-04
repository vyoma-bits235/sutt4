from django.shortcuts import render,HttpResponse,redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth import authenticate,login,logout
from django.contrib.auth.decorators import login_required,user_passes_test
from django.shortcuts import render,HttpResponse,redirect
from .models import MenuItem, Rating, Orders,Image23,Attendance,AT2,Menu,ExcelData,ExcelFile,Expenditure,Items,UserProfile
from polls.models import Menu


from django.utils import formats
from social_django.models import UserSocialAuth
from datetime import datetime, timedelta
from math import ceil
# import the logging library
import logging
from django.db.models import Avg
import pandas 
import xlrd
from xlrd import open_workbook
import datetime
from datetime import datetime
from datetime import timezone
import pandas as pd
import openpyxl
from django.core.files.storage import FileSystemStorage


from PIL import Image


now = datetime.now()


# Get the time zone information


# Use the time zone information to create a datetime object with the specified time zone

def is_student(user):
    return UserProfile.objects.get(user=user).role == 'student'

def is_staff(user):
    return UserProfile.objects.get(user=user).role == 'staff'

def login(request):
    return render(request,'login.html')

def h1(request):
    return render(request,'basic.html')

def home(request):
        current_time = datetime.now()

    # Identify the current student
    

    # Determine the upcoming meal
        next_meal = None
        if current_time.hour < 11:
            next_meal = 'breakfast'
        elif current_time.hour < 14:
            next_meal = 'lunch'
        elif current_time.hour <20:
            next_meal = 'dinner'
        else:
            next_meal = 'breakfastn1'
        
        today = datetime.today()
        
        
# Shift the date to the next day
        
        next_meal_menu=None
        tomorrow=today+timedelta(days=1)
        a=today.strftime("%Y-%m-%d")+" "
        b=tomorrow.strftime("%Y-%m-%d")+" "
    
       
        
        
        
        
        

        if next_meal == 'breakfast':
            next_meal_menu = Items.objects.get(date=a)
        elif next_meal == 'lunch':
            next_meal_menu = Items.objects.get(date=a)
        elif next_meal == 'dinner':
            next_meal_menu =Items.objects.get(date=a)
        elif next_meal == 'breakfastn1':
            next_meal_menu =Items.objects.get(date=b)
        a={}
        a["next_meal_menu"]=next_meal_menu
        a["next_meal"]=next_meal

        

        
        return render(request,'index.html',a)

def SignupPage(request):
    if request.method=='POST':
        uname=request.POST.get('username')
        email=request.POST.get('email')
        pass1=request.POST.get('password1')
        pass2=request.POST.get('password2')

        if pass1!=pass2:
            return HttpResponse("Your password and confrom password are not Same!!")
        else:

            my_user=User.objects.create_user(uname,email,pass1)
            my_user.save()
            return redirect('login1')
        



    return render (request,'signup.html')

def LoginPage(request):
    if request.method=='POST':
        username=request.POST.get('username')
        pass1=request.POST.get('pass')
        user=authenticate(request,username=username,password=pass1)
        if user is not None:
            return redirect('downljhuodad')
 
        else:
            return HttpResponse ("Username or Password is incorrect!!!")

    return render (request,'login1.html')
def staff_dashboard(request):
    feedbacks = Image23.objects.all()


    context = {
        'feedbacks': feedbacks,
        
    }
    return render(request, 'st_dash2.html', context)
def che(l):
            
            c=0
            for i in Rating.objects.all():
                    if l==i.menu_item:
                        c=c+1
            if c==0:
                ar=Rating(menu_item=l)
                ar.save()
            
def menu_display(request):
    menu_items = Items.objects.all()
    if request.method == 'POST':
        r1=request.POST.get('r1')
        r2=request.POST.get('r2')
        r3=request.POST.get('r3')
        r4=request.POST.get('r4')
        r5=request.POST.get('r5')
        r6=request.POST.get('r6')
        r7=request.POST.get('r7')
        r8=request.POST.get('r8')
        r9=request.POST.get('r9')
        r10=request.POST.get('r10')
        r11=request.POST.get('r11')
        r12=request.POST.get('r12')
        r13=request.POST.get('r13')
        r14=request.POST.get('r14')
        r15=request.POST.get('r15')
        r16=request.POST.get('r16')
        r17=request.POST.get('r17')
        r18=request.POST.get('r18')
        r19=request.POST.get('r19')
        r20=request.POST.get('r20')
        r21=request.POST.get('r21')
        r22=request.POST.get('r22')
        r23=request.POST.get('r23')
        r24=request.POST.get('r24')
        l=[r1,r2,r3,r4,r5,r6,r7,r8,r9,r10,r11,r12,r13,r14,r15,r16,r17,r18,r19,r20,r21,r22,r23,r24]
        item1=request.POST.get('item1')
        item2=request.POST.get('item2')
        item3=request.POST.get('item3')
        item4=request.POST.get('item4')
        item5=request.POST.get('item5')
        item6=request.POST.get('item6')
        item7=request.POST.get('item7')
        item8=request.POST.get('item8')
        item9=request.POST.get('item9')
        item10=request.POST.get('item10')
        item11=request.POST.get('item11')
        item12=request.POST.get('item12')
        item13=request.POST.get('item13')
        item14=request.POST.get('item14')
        item15=request.POST.get('item15')
        item16=request.POST.get('item16')
        item17=request.POST.get('item17')
        item18=request.POST.get('item18')
        item19=request.POST.get('item19')
        item20=request.POST.get('item20')
        item21=request.POST.get('item21')
        item22=request.POST.get('item22')
        item23=request.POST.get('item23')
        item24=request.POST.get('item24')
        iu=[item1,item2,item3,item4,item5,item6,item7,item8,item9,item10,item11,item12,item13,item14,item15,item16,item17,item18,item19,item20,item21,item22,item23,item24]
        che(item1)
        che(item2)
        che(item3)
        che(item4)
        che(item5)
        che(item6)
        che(item7)
        che(item8)
        che(item9)
        che(item10)
        che(item11)
        che(item12)
        che(item13)
        che(item14)
        che(item15)
        che(item16)
        che(item17)
        che(item18)
        che(item19)
        che(item20)
        che(item21)
        che(item22)
        che(item23)
        che(item24)
        j=0
        for i in Rating.objects.all():
            i.rating=i.rating+int(l[j])
            i.person=i.person+1
            i.avg=i.rating/i.person
            i.save()
            j=j+1




       
        
        
    return render(request, 'menu_display.html', {'menu_items': menu_items})

def rating1(request):
   
    if request.method == 'POST':
        address2=request.POST.get('username','')
        

        
        
       


        return render(request,"rating.html")
   




def checkout(request):
    if request.method=="POST":
        
      
        name = request.POST.get('name', '')
       
        address = request.POST.get('address1', '') + " " + request.POST.get('address2', '')
        city = request.POST.get('city', '')
       
        order = Image23(content=city)
        order.save()
        thank = True
        id = order.order_id
        
    return render(request, 'checkout.html')
def mark_presence(request):
    ab=""
    current_time = datetime.now()
    mess_timings = [
            (datetime.strptime('7:00 AM', '%I:%M %p'), datetime.strptime('9:00 AM', '%I:%M %p')),
            (datetime.strptime('12:00 PM', '%I:%M %p'), datetime.strptime('2:00 PM', '%I:%M %p')),
            (datetime.strptime('7:00 PM', '%I:%M %p'), datetime.strptime('9:00 PM', '%I:%M %p')),
        ]
    if datetime.strptime('7:00 AM', '%I:%M %p').strftime('%I:%M %p') <= current_time.strftime('%I:%M %p') <= datetime.strptime('9:00 AM', '%I:%M %p').strftime('%I:%M %p'):
            ab="breakfast"
    elif datetime.strptime('12:00 PM', '%I:%M %p').strftime('%I:%M %p') <= current_time.strftime('%I:%M %p') <= datetime.strptime('2:00 PM', '%I:%M %p').strftime('%I:%M %p'):
            ab="lunch"
    elif  datetime.strptime('7:00 PM', '%I:%M %p').strftime('%I:%M %p') <= current_time.strftime('%I:%M %p') <= datetime.strptime('9:00 PM', '%I:%M %p').strftime('%I:%M %p'):
            ab="dinner"
    else:
        ab="Not allowed"
    if request.method=="POST":
        current_time = datetime.now()
        meal_time=request.POST.get("meal_time")


        mess_timings = [
            (datetime.strptime('7:00 AM', '%I:%M %p'), datetime.strptime('9:00 AM', '%I:%M %p')),
            (datetime.strptime('12:00 PM', '%I:%M %p'), datetime.strptime('2:00 PM', '%I:%M %p')),
            (datetime.strptime('7:00 PM', '%I:%M %p'), datetime.strptime('9:00 PM', '%I:%M %p')),
        ]
        ab=""
        is_within_mess_hours = False
        for start_time, end_time in mess_timings:
    
            if start_time.strftime('%I:%M %p') <= current_time.strftime('%I:%M %p') <= end_time.strftime('%I:%M %p'):
                is_within_mess_hours = True
                break
        if datetime.strptime('7:00 AM', '%I:%M %p').strftime('%I:%M %p') <= current_time.strftime('%I:%M %p') <= datetime.strptime('9:00 AM', '%I:%M %p').strftime('%I:%M %p'):
            ab="breakfast"
        elif datetime.strptime('12:00 PM', '%I:%M %p').strftime('%I:%M %p') <= current_time.strftime('%I:%M %p') <= datetime.strptime('2:00 PM', '%I:%M %p').strftime('%I:%M %p'):
            ab="lunch"
        else:
            ab="dinner"
       
            
            
        

        if not is_within_mess_hours:
            # Display an error message if the student is not marking during mess hours
            error_message = 'Attendance marking is only allowed during mess hours.'
            context = {'error_message': error_message}
            return render(request, 'failed.html', context)

        # Check if the student has already marked their attendance for the current meal
        # (Use student's attendance records to determine this)

        # If the student has not yet marked their attendance for the current meal,
        # mark their attendance and redirect them to the dashboard with a success message

        success_message = 'Attendance marked successfully.'
        stu=request.POST.get('use1')
        
        c=0
        for i in Attendance.objects.all():
                    if stu==i.student:
                        c=c+1
        if c==0:
            ar=Attendance(student=stu)
            ar.save()
        de=0
        for j in AT2.objects.all():
            if datetime.today().strftime("%Y-%m-%d") == j.date:
                de=de+1
        if de==0:
            arp=AT2(date=datetime.today().strftime("%Y-%m-%d"))
            arp.save()
      

        w=AT2.objects.get(date=datetime.today().strftime("%Y-%m-%d"))
        if meal_time=="breakfast":
            w.breakfast=int(w.breakfast)+1
        if meal_time=="lunch":
            w.lunch=int(w.lunch)+1
        if meal_time=="dinner":
            w.Dinner=int(w.Dinner)+1
        w.save()
        m="ab"

           

     
                        
                    
                    
        


        
        meal_ti=request.POST.get('meal_time')
        att = Attendance.objects.get(student="admin")
        if meal_ti=="breakfast":
            att.breakfast=int(att.breakfast)+1
        if meal_ti=="lunch":
            att.lunch=int(att.lunch)+1
        if meal_ti=="dinner":
            att.Dinner=int(att.Dinner)+1
        att.save()

        

        

        context = {'success_message': success_message}
        return render(request,"success.html",context)
    

    return render(request, 'mark_presence.html',{"m":ab})
def calculate_fees(request):
    # Get the attendance data for the student
    d1=Attendance.objects.all()

    k=0
    for i in d1:
        c=0
        for j in Expenditure.objects.all():
                    if i.student==j.name:
                        c=c+1
        if c==0:

           

            d2=Expenditure(name=i.student)
            d2.save()
        u=i.student
        ci=Expenditure()
        attendance_data = get_attendance_data(u)
        total_cost_breakfast =attendance_data[0] * 80
        total_cost_lunch = attendance_data[1] * 150
        total_cost_dinner =attendance_data[2] * 80


    # Calculate the total fees
        total_fees = total_cost_breakfast + total_cost_lunch + total_cost_dinner
        k=total_fees
        ty=Expenditure.objects.get(name=u)
        ty.exp=total_fees
        ty.save()


    
    
   
    

    
    

   
   


     #Calculate the total cost of breakfast, lunch, and dinner
    


    # Render the template with the student's fees
    return HttpResponse(k)

# Get the attendance data for the student
def get_attendance_data(student_id):
    # Get the attendance data from the database
    attendance_data = Attendance.objects.get(student=student_id)
    o=[]
    if len(attendance_data.breakfast)>=0:

        o.append(int(attendance_data.breakfast))
        o.append(int(attendance_data.lunch))
        o.append(int(attendance_data.Dinner))
    
    return o
def count(request):
    ad=AT2.objects.all()
    meal={}
    b1=0
    l1=0
    d1=0
    for i in ad:
        b1=b1+int(i.breakfast)
        l1=l1+int(i.lunch)
        d1=d1+int(i.Dinner)
    meal["breakfast"]=b1
    meal["lunch"]=l1
    meal["dinner"]=d1
    return render(request,'count.html',meal)
def get_next_meal_menu(request):
    # Retrieve the current time
    current_time = datetime.now()

    # Identify the current student
    

    # Determine the upcoming meal
    next_meal = None
    if current_time.hour < 11:
        next_meal = 'breakfast'
    elif current_time.hour < 14:
        next_meal = 'lunch'
    elif current_time.hour <20:
        next_meal = 'dinner'
    else:
        next_meal = 'breakfastn1'
    next_meal_menu = None
    today = datetime.today()
# Shift the date to the next day
    tomorrow = today - timedelta(days=1)
    if next_meal == 'breakfast':
        next_meal_menu = Menu.objects.get(date=formats.date_format(today, Menu._meta.get_field('date_field').format)).breakfast
    elif next_meal == 'lunch':
        next_meal_menu = Menu.objects.get(date=formats.date_format(today, Menu._meta.get_field('date_field').format)).lunch
    elif next_meal == 'dinner':
        next_meal_menu = Menu.objects.get(date=formats.date_format(today, Menu._meta.get_field('date_field').format)).dinner
    elif next_meal=="breakfastn1":
        next_meal_menu=Menu.objects.get(date=formats.date_format(tomorrow, Menu._meta.get_field('date_field').format)).breakfast
    return HttpResponse("next_meal_menu")


    return render(request,'index.html',{"next_meal_menu":next_meal_menu})
def remove(l):# function for removing empty spaces and * pattern
    p=[]
    for i in l:
        if( str(i) in "*****************" or str(i)=="nan"):
            p.append(".")
        else:
            p.append(i)
    return p




def pass_file(file):  # function for parsing the excel file
    df = pd.read_excel(file,header=1)
    l=[]
    a=[]
   

    for col in df.columns:
        
        a.append(col)
  
    for j in range(15):
        meals=[]

        for i, row in df.iterrows():
        
            meal = row[a[j]]
            meals.append(meal)
            
        l.append(meals)
   
    w=[]
    for y in range(15):# logic for the loop is that we slice the list till the contents of dinner,lunch,breakfast and then we remove the spaes or **** pattern
        
    
        
        o=l[y][22:29][:]
        d=remove(o)
        
        
        r=l[y][1:10][:]
        b=remove(r)
        
        c=l[y][12:20][:]
        t=remove(c)
        
       
            
      
       
        f=[{"Breakfast":b,"Lunch":t,"Dinner":d}]
       
        e=[str(a[y])[0:11]]
        e.append(f)
        w.append(e)
    return w
def extract(l):
    a=""
    for i in l:
        a=a+" "+i
    return a
  

def menu_upload(request):
    file = 'static/Mess Menu.xlsx'
    meals = pass_file(file)
    for i in range(len(meals)):
        a=meals[i][0]
        a.strip()



    
        mk=Menu(date=a,breakfast=extract(meals[i][1][0].get('Breakfast')),lunch=extract(meals[i][1][0].get('Lunch')),dinner=extract(meals[i][1][0].get('Dinner')))
        mk.save()
    return redirect("downljhuodad")
def google_name(request):
    
    if True:
        # Get the current user
        user1 =request.POST.get('user')
        # Retrieve the Google name from the UserSocialAuth object
        user_social_auth = UserSocialAuth.objects.get(user=14)
        google_name = user_social_auth.extra_data['auth_time']
        return HttpResponse(google_name)
    else:
        return HttpResponse("madarchod")
def upload_image(request):
    if request.method == 'POST':
        # Get the image file from the request
        image_file = request.FILES['image']
        nam1e = request.POST['name']
       


        # Open the image file as an Image object
        
        a=Image23(image=image_file,content=nam1e)
        a.save()

        # Return the image data as an HTTP response
        return HttpResponse("jnjn")
    else:
        # Render the upload form HTML
        return render(request, 'upload_image.html')
def save_excel_file_data(excel_file):
    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file.file)

    # Get the first worksheet in the workbook
    worksheet = workbook.worksheets[0]

    # Iterate through the rows and columns of the worksheet
    for row in worksheet.iter_rows(min_row=2):
        # Create a new instance of the ExcelData model
        excel_data = ExcelData()

        # Set the data fields based on the values in the current row
        excel_data.name = row[0].value
        excel_data.value = row[1].value

        # Save the ExcelData instance to the database
        excel_data.save()
def upload_excel_file(request):
    if request.method == 'POST':
        # Get the uploaded Excel file
        excel_file = request.FILES['excel_file']

        # Save the Excel file to the media directory using FileSystemStorage
        fs = FileSystemStorage()
        filename = fs.save('static/' + excel_file.name, excel_file)

        # Create an ExcelFile instance with the saved filename
        saved_excel_file = ExcelFile(file=filename)

        # Save the ExcelFile instance to the database
        saved_excel_file.save()

        # Save the Excel file data to the database
        

        # Return a success message
        return HttpResponse('Excel file uploaded and processed successfully.')
    else:
        # Render the upload form HTML
        return render(request, 'upload_excel.html')
def data3(request):
    return render(request,"download_data.html")

from django.http import HttpResponse
from django.shortcuts import render
import xlsxwriter
from django.http import FileResponse
from django.conf import settings
import os

def data(request):
   
    file=os.path.join(settings.BASE_DIR,'Expenditure.xlsx')
    fileopened=open(file,'rb')
    return FileResponse(fileopened)
def data2(request):
   
    file=os.path.join(settings.BASE_DIR,'Expenditure.xlsx')
    fileopened=open(file,'rb')
    return FileResponse(fileopened)

def data1(request):
    return render(request,"data1.html")
def hi(request):
    file=os.path.join(settings.BASE_DIR,'Expenditure.xlsx')
    fileopened=open(file,'rb')
    return FileResponse(fileopened)
def index1(request):
    return render(request,"index.html")
def start(request):
    return render(request,"staff2.html")


def expenditure(request):
    ER=Expenditure.objects.all()
    a=0
    h=0
    for i in ER:
        h=h+1
        a=int(i.exp)+a
    
    FT = AT2.objects.all()


    FTE=Rating.objects.all()
    
    for i in FTE:
        i.avg=i.rating/i.person
        i.save()
        
    
   

   
    

    
    

    return render(request,"staff2.html",{"ER":ER,"FT":FT,"a":a,"h":h,"FTE":FTE})
def time(request):
    FT=AT2.objects.all()
    return render(request,"staff2.html",{"FT":FT})
def ju(request):
        file=os.path.join(settings.BASE_DIR,'WIN_20231025_23_08_19_Pro_wKP65t0.jpg')
        fileopened=open(file,'rb')
        return FileResponse(fileopened)
def uio(request):
    return  render(request,"uio.html")
def profile(request):
    full_name = request.user.get_full_name()
    
    
    return  render(request,"profile.html",{"name":full_name})
def send(request):


# Load the Excel file
    wb = openpyxl.load_workbook('Expenditure.xlsx')

# Get the worksheet you want to update
    ws = wb.active

# Get the next row number
    next_row = ws.max_row + 1
    gt=Expenditure.objects.all()
    for i in gt:

# Add a new row of data
        name = i.name
        exp1=i.exp



    ws.append([name,exp1])

# Save the changes to the Excel file
    wb.save('Expenditure.xlsx')
    return HttpResponse("dhdj")
def menu_up1(request):
    file = 'static/Mess Menu.xlsx'
    meals = pass_file(file)
    for i in range(len(meals)):
        a=meals[i][0]
        a.strip()
        
        
        
        item=Items(date=a)
        item.save()
        it1=Items.objects.get(date=a)
        it1.item1=meals[i][1][0].get('Breakfast')[0]
        it1.item2=meals[i][1][0].get('Breakfast')[1]
        it1.item3=meals[i][1][0].get('Breakfast')[2]
        it1.item4=meals[i][1][0].get('Breakfast')[3]
        it1.item5=meals[i][1][0].get('Breakfast')[4]
        it1.item6=meals[i][1][0].get('Breakfast')[5]
        it1.item7=meals[i][1][0].get('Breakfast')[6]
        it1.item8=meals[i][1][0].get('Breakfast')[7]
        it1.item9=meals[i][1][0].get('Breakfast')[8]
        it1.item10=meals[i][1][0].get('Lunch')[0]
        it1.item11=meals[i][1][0].get('Lunch')[1]
        it1.item12=meals[i][1][0].get('Lunch')[2]
        it1.item13=meals[i][1][0].get('Lunch')[3]
        it1.item14=meals[i][1][0].get('Lunch')[4]
        it1.item15=meals[i][1][0].get('Lunch')[5]
        it1.item16=meals[i][1][0].get('Lunch')[6]
        it1.item17=meals[i][1][0].get('Lunch')[7]
        it1.item18=meals[i][1][0].get('Dinner')[0]
        it1.item19=meals[i][1][0].get('Dinner')[1]
        it1.item20=meals[i][1][0].get('Dinner')[2]
        it1.item21=meals[i][1][0].get('Dinner')[3]
        it1.item22=meals[i][1][0].get('Dinner')[4]
        it1.item23=meals[i][1][0].get('Dinner')[5]
        it1.item24=meals[i][1][0].get('Dinner')[6]
        it1.save()
    return HttpResponse("OKJ")
@login_required
def view_name(request):
    # This will get the username
    username = request.user.username

    # This will get the full name (i.e., first name and last name)
    full_name = request.user.get_full_name()

    return HttpResponse(f'Hello, {full_name} ({username})')





    
    

    
    








        
    

